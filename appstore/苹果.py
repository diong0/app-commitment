import re
import requests
import json
import time
import openpyxl
from datetime import datetime
import jieba

def load_stopwords(filepath):
    with open(filepath, 'r', encoding='utf-8') as file:
        stopwords = set(line.strip() for line in file)
    return stopwords

def remove_stopwords(words, stopwords):
    return [word for word in words if word not in stopwords]

def search_app_id(app):
    url = "http://itunes.apple.com/search?term=" + app + "&entity=software"
    r = requests.get(url)
    html = r.content
    html_doc = str(html, 'utf-8')
    data = json.loads(html_doc)
    resultCount = data['resultCount']
    results = data['results']
    print(app + " Find " + str(resultCount) + " result(s)")
    if resultCount > 0:
        name = results[0]['trackName']
        app_id = results[0]['trackId']
        print("Defaulting to first result: name：" + name, "id：" + str(app_id))
        return app_id
    else:
        print("No results found for " + app)
        return None

def save_content(id, wb, ws, stopwords, comment_data, appname):
    row = 2
    total_rating = 0
    max_length = 100
    for j in range(1, 11):
        url = "https://itunes.apple.com/rss/customerreviews/page=" + str(j) + "/id=" + str(id) + "/sortby=mostrecent/json?l=en&&cc=cn"
        r = requests.get(url)
        if r.status_code == 200:
            html = r.content
            html_doc = str(html, 'utf-8')
            data = json.loads(html_doc).get("feed", {}).get("entry", [])
            for i in data:
                name = i['author']['name']['label']
                rate = i['im:rating']['label']
                content = i['content']['label']
                updated = i['updated']['label']
                version = i['im:version']['label']

                updated_date = datetime.strptime(updated, "%Y-%m-%dT%H:%M:%S%z")
                formatted_date = updated_date.strftime("%Y/%m/%d")

                total_rating += int(rate)
                re_content = re.sub(r'[^\w\s]', '', content)
                re_content = re_content.replace('\n', '')
                re_content = re_content.replace(' ', '')

                if len(content) > max_length:
                    content = content[:max_length] + "..."
                text = jieba.lcut(re_content)
                filtered_text = remove_stopwords(text, stopwords)

                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=content)
                ws.cell(row=row, column=3, value=rate)
                ws.cell(row=row, column=4, value=formatted_date)
                ws.cell(row=row, column=5, value=version)
                row += 1
                comment_data.append([str(name), appname, formatted_date, '华为', version, content, rate, [filtered_text]])
        else:
            return
        time.sleep(2)
    ws.cell(row=1, column=6, value="总评分")
    ws.cell(row=2, column=6, value=total_rating / 500)

def fetch_comments(appname, stopwords_filepath='stop.txt'):
    stopwords = load_stopwords(stopwords_filepath)
    comment_data = []
    app_id = search_app_id(appname)
    if app_id is None:
        print("Exiting due to no results found.")
        return comment_data

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="评论ID")
    ws.cell(row=1, column=2, value="评论内容")
    ws.cell(row=1, column=3, value="评论评分")
    ws.cell(row=1, column=4, value="评论时间")
    ws.cell(row=1, column=5, value="评论版本")

    save_content(app_id, wb, ws, stopwords, comment_data, appname)

    wb.save(appname + ".xlsx")
    print("Done!")
    return comment_data
