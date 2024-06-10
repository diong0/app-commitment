#encoding=utf-8
#!/usr/bin/env

import re
import requests
import json
import time
import openpyxl
from datetime import datetime
import jieba


def SearchAppId(app):
    url = "http://itunes.apple.com/search?term=" + app + "&entity=software"
    r = requests.get(url)
    html = r.content
    html_doc = str(html, 'utf-8')
    data = json.loads(html_doc)
    resultCount = data['resultCount']
    results = data['results']
    print(app + " Find " + str(resultCount) + " result(s)")
    if resultCount > 0:
        name = results[0]['trackName']
        app_id = results[0]['trackId']
        print("Defaulting to first result: name：" + name, "id：" + str(app_id))
        return app_id
    else:
        print("No results found for " + app)
        return None


# 读取停用词表
def load_stopwords(filepath):
    with open(filepath, 'r', encoding='utf-8') as file:
        stopwords = set(line.strip() for line in file)
    return stopwords


# 去除停用词
def remove_stopwords(words, stopwords):
    return [word for word in words if word not in stopwords]


# 加载停用词表
stopwords = load_stopwords('../stop.txt')


def SaveContent(id, wb, ws):
    row = 2
    total_rating = 0
    max_length = 100  # 评论内容最大长度
    for j in range(1, 11):  # 只能爬取前十页
        url = "https://itunes.apple.com/rss/customerreviews/page=" + str(j) + "/id=" + str(
            id) + "/sortby=mostrecent/json?l=en&&cc=cn"
        r = requests.get(url)

        if r.status_code == 200:
            html = r.content
            html_doc = str(html, 'utf-8')
            data = json.loads(html_doc).get("feed", {}).get("entry", [])
            for i in data:
                name = i['author']['name']['label']
                rate = i['im:rating']['label']
                content = i['content']['label']
                updated = i['updated']['label']
                version = i['im:version']['label']

                # 转换时间格式
                updated_date = datetime.strptime(updated, "%Y-%m-%dT%H:%M:%S%z")
                formatted_date = updated_date.strftime("%Y/%m/%d")

                # 累加总评分
                total_rating += int(rate)
                # 去除符号
                re_content = re.sub(r'[^\w\s]', '', content)
                re_content = re_content.replace('\n', '')
                # 截断评论内容
                if len(content) > max_length:
                    content = content[:max_length] + "..."
                # 分词
                text = jieba.lcut(re_content)

                # 去除停用词
                filtered_text = remove_stopwords(text, stopwords)

                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=content)
                ws.cell(row=row, column=3, value=rate)
                ws.cell(row=row, column=4, value=formatted_date)
                ws.cell(row=row, column=5, value=version)
                row = row + 1
                print(name, rate, content, formatted_date, version)

        else:
            return
        # 每一页爬取延迟2秒，以防过于频繁
        time.sleep(2)
    # 在Excel表格中添加总评分
    ws.cell(row=1, column=6, value="总评分")
    ws.cell(row=2, column=6, value=total_rating / 500)


def main():
    app = input("app:\n")
    app_id = SearchAppId(app)

    if app_id is None:
        print("Exiting due to no results found.")
        return

    # Workbook init
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="评论ID")
    ws.cell(row=1, column=2, value="评论内容")
    ws.cell(row=1, column=3, value="评论评分")
    ws.cell(row=1, column=4, value="评论时间")
    ws.cell(row=1, column=5, value="评论版本")

    SaveContent(app_id, wb, ws)

    wb.save(app + ".xlsx")  # 默认保存在当前目录
    print("Done!")


if __name__ == '__main__':
    main()
