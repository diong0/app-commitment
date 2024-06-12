import openpyxl
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from time import sleep
from lxml import etree  # 导入lxml库中的etree模块
from datetime import datetime
import re
import jieba


def scrape_app_comments(appname):
    # 设置ChromeDriver的路径
    chromedriver_path = r"D:\anaconda\chromedriver.exe"

    # 创建ChromeDriver服务对象
    service = Service(chromedriver_path)

    # 创建Chrome浏览器实例
    driver = webdriver.Chrome(service=service)

    # 打开一个网页
    driver.get("https://appgallery.huawei.com/#/Apps")

    # 在浏览器执行JavaScript代码，用于欺骗检测机制，隐藏webdriver属性
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
                Object.defineProperty(navigator, 'webdriver', {
                  get: () => undefined
                })
              """
    })
    # 等待页面加载
    driver.implicitly_wait(10)  # 隐式等待，最多等10秒

    driver.find_element(By.XPATH, '//*[@id="app"]/div/div[1]/div[3]/div[1]/input').send_keys(appname)
    driver.find_element(By.XPATH, '//*[@id="app"]/div/div[1]/div[3]/div[1]/input').send_keys(Keys.ENTER)
    sleep(2)
    driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[1]/div/div[1]').click()
    sleep(2)
    driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[3]/div[4]/div/div[1]/div[2]/span').click()

    # 循环爬取评论的页数，每页25条评论
    for i in range(0, 4):
        # 执行JavaScript代码，将页面滚动到底部
        driver.execute_script('window.scrollTo(0,document.body.scrollHeight)')
        # 等待1秒
        sleep(1)
    # 获取页面的源代码
    html = driver.page_source

    # 使用lxml解析页面源代码
    tree = etree.HTML(html)

    # 获取评论列表
    comment_list = tree.xpath('//*[@id="app"]/div/div[2]/div/div[3]/div[11]/div[4]/div')

    # 创建一个新的Excel工作簿
    workbook = openpyxl.Workbook()
    # 选择默认的工作表
    sheet = workbook.active
    # 写入表头
    sheet["A1"] = "评论ID"
    sheet["B1"] = "评论内容"
    sheet["C1"] = "评论评分"
    sheet["D1"] = "评论时间"
    sheet["E1"] = "评论版本"
    sheet["F1"] = "总评分"
    # 定义一个空列表来存储评论数据
    comment_data = []

    # 获取应用总评分
    score_element = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[3]/div[11]/div[3]/div[1]/div[1]')
    score = score_element.text

    # 获取评分总人数
    # score_number_element = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[3]/div[11]/div[3]/div[1]/div[2]')
    # score_number = score_number_element.text.replace("人评分", "").strip()
    # 获取各个评分人数

    # 读取停用词表
    def load_stopwords(filepath):
        with open(filepath, 'r', encoding='utf-8') as file:
            stopwords = set(line.strip() for line in file)
        return stopwords

    # 去除停用词
    def remove_stopwords(words, stopwords):
        return [word for word in words if word not in stopwords]

    # 加载停用词表
    stopwords = load_stopwords('stop.txt')

    # 遍历评论列表,最多100次
    count = 0
    # 评论最大长度
    max_length = 100
    for div in comment_list:
        if count >= 100:
            break
        # 获取评论者的昵称
        name = div.xpath('.//div[@class="userName"]/text()')[0]
        # 获取评论的时间
        raw_time = div.xpath('./div[2]/div[1]/div/div[2]/div/text()')[0]
        parsed_time = datetime.strptime(raw_time, "%m/%d/%Y, %H:%M")
        time = parsed_time.strftime("%Y/%m/%d")
        # 获取评论的内容
        comment = div.xpath('.//div[@class="part_middle"]/text()')[0]

        # 去除符号
        re_comment = re.sub(r'[^\w\s]', '', comment)
        re_comment = re_comment.replace('\n', '')
        re_comment = re_comment.replace(' ', '')
        # 截断评论内容
        if len(comment) > max_length:
            comment = comment[:max_length] + "..."
        # 分词
        text = jieba.lcut(re_comment)

        # 去除停用词得到最终分词
        filtered_text = remove_stopwords(text, stopwords)
        # 重新连接为字符串
        filtered_comment = ' '.join(filtered_text)

        # 获取单评论的评分
        star_elements = div.xpath('.//div[@class="newStarBox starBox"]/img')
        rating = 0
        for star in star_elements:
            star_src = star.get('src')
            if '8L3N2Zz4=' in star_src:
                rating += 1  # 包含8L3N2Zz4=编码的图片表示黄色星星

        # 获取应用版本号
        raw_user_version = div.xpath('.//div[@class ="version"]/text()')[0]
        user_version = raw_user_version.strip("()")
        # 将评论数据添加到列表中
        comment_data.append([str(name), appname, time, '华为', user_version, comment, rating, [filtered_comment]])
        count += 1
    # 循环写入评论数据到Excel
    for i, comment in enumerate(comment_data):
        # 计算行号
        row = i + 2
        # 写入评论数据到Excel单元格
        sheet.cell(row=row, column=1, value=comment[0])
        sheet.cell(row=row, column=2, value=comment[1])
        sheet.cell(row=row, column=3, value=comment[2])
        sheet.cell(row=row, column=4, value=comment[3])
        sheet.cell(row=row, column=5, value=comment[4])
        sheet.cell(row=row, column=6, value=comment[5])
        # 保存工作簿
        workbook.save("Huawei.xlsx")
    # 关闭浏览器
    driver.quit()

    return comment_data
    #   评论id appid 评论时间 应用商店id  评论版本  评论内容 评论评分 评论分词
